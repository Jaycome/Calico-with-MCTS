from copy import deepcopy
import numpy as np
from meta import GameMeta
import random
from colorama import Fore, Back, Style
from mcts import UctMctsAgent
from mcts import Node
from mcts2 import UctMctsAgent2
from ConnectState import GameState
# import xlsxwriter module
import xlsxwriter
import openpyxl
if __name__ == "__main__":
    roundTime=50
    searchTime=0.5
    player1Win=0
    player2Win=0
    tie=0
    buttonToken=0
    catToken=0
    designGoal=0
    total=0
    buttonToken2=0
    catToken2=0
    designGoal2=0
    total2=0
    wholeTotal=[]
    wholeTotal2=[]

    #random vs. random    
    """wb = openpyxl.load_workbook('MCTSVsRandom.xlsx', data_only=True)

    s1 = wb.create_sheet('worksheet 1')     # add worksheet 1
    data = []
    dataName=['','player1','player2']
    data.append(dataName)
    for i in range(roundTime):
        print("round",i+1,'...')
        state=GameState()
        state.initial_setting()
        state.cat_setting()
        state.set_goal_tile(state.tile_base)
        state.set_goal_tile(state.tile_base2)
        state.place_goal_tile(state.tile_base,(0,1,2),(6,13,16))
        state.place_goal_tile(state.tile_base2,(0,1,2),(6,13,16))
        
        roundCount=[]
        roundCount.append('round'+str(i+1))
        while state.round<=22:
            moves=state.possible_moves()
            move=random.choice(moves)
            state.move(move)
            #roundCount.append(len(count2))
            moves=state.possible_moves2()
            move=random.choice(moves)
            state.move(move)
        pointsFromElement=0
        pointsFromPattern=0
        pointsFromElement+=state.rainbow_button(state.tile_base)
        for i in range(len(state.tile_base)):
            if(state.tile_base[i]['token']==1):
                pointsFromElement+=3
            elif(state.tile_base[i]['token']==2):
                pointsFromPattern+=3
            elif(state.tile_base[i]['token']==3):
                pointsFromPattern+=5
            elif(state.tile_base[i]['token']==4):
                pointsFromPattern+=7
        pointsFromGoal=state.calculate_goal_points(state.tile_base,(6,13,16))
        pointsTotal=pointsFromPattern+pointsFromElement+pointsFromGoal
        roundCount.append(pointsTotal)
 
        points2FromElement=0
        points2FromPattern=0
        points2FromElement+=state.rainbow_button(state.tile_base2)
        for i in range(len(state.tile_base2)):
            if(state.tile_base2[i]['token']==1):
                points2FromElement+=3
            elif(state.tile_base2[i]['token']==2):
                points2FromPattern+=3
            elif(state.tile_base2[i]['token']==3):
                points2FromPattern+=5
            elif(state.tile_base[i]['token']==4):
                points2FromPattern+=7
        points2FromGoal=state.calculate_goal_points(state.tile_base2,(6,13,16))
        points2Total=points2FromPattern+points2FromElement+points2FromGoal
        roundCount.append(points2Total)
        if(state.winner==1):
            player1Win+=1
        elif(state.winner==2):
            player2Win+=1
        else:
            tie+=1
        data.append(roundCount)
    for i in data:
        s1.append(i)
    wb.save('t_test.xlsx')"""
#MCTS vs. MCTS

    workbook = xlsxwriter.Workbook('MCTS4VsMCTS05.xlsx')
    wb = openpyxl.load_workbook('MCTS4VsMCTS05.xlsx', data_only=True)

    s1 = wb.create_sheet('worksheet1')     # add worksheet 1
    s2 = wb.create_sheet('worksheet2')     # add worksheet 2
    data = []
    data2 = []
    dataName=['','MCTS_0','','','','MCTS_1']
    data.append(dataName)
    dataName=['','button token','cat token','goal tile','total points','button token','cat token','goal tile','total points']
    data.append(dataName)
    data2Name=[]
    data2Name.append('')
    for i in range(21):
        data2Name.append('turn'+str(i+1))
        data2Name.append('')
    data2.append(data2Name)

    data2Name=['']
    for i in range(21):
        data2Name.append('MCTS_0')
        data2Name.append('MCTS_1')
    data2.append(data2Name)
    data2Name=[]



    for i in range(roundTime):
        print("round",i+1,'...')
        state=GameState()
        state.initial_setting()
        state.cat_setting()
        state.set_goal_tile(state.tile_base)
        state.set_goal_tile(state.tile_base2)
        state.place_goal_tile(state.tile_base,(0,1,2),(6,13,16))
        state.place_goal_tile(state.tile_base2,(0,1,2),(6,13,16))

        mcts=UctMctsAgent(state)
        mcts2=UctMctsAgent2(state)

        
        rollOut=[]
        roundCount=[]
        roundCount.append('round'+str(i+1))
        rollOut.append('round'+str(i+1))
        while state.round<=22:
            if(state.round==22):
                biggest=[]
                moves=state.possible_moves()
                move=[]
                for i in moves:
                    pointsFromElement=state.calculate_element_points(state.tile_base)
                    pointsFromPattern=state.calculate_pattern_points(state.tile_base)
                    pointsFromGoal=state.calculate_goal_points(state.tile_base,(6,13,16))
                    pointsTotal=pointsFromPattern+pointsFromElement+pointsFromGoal
                    move=list(i)
                    state.last_player1_place_tile(move)
                    pointsFromElement=state.calculate_element_points(state.tile_base)
                    pointsFromPattern=state.calculate_pattern_points(state.tile_base)
                    pointsFromGoal=state.calculate_goal_points(state.tile_base,(6,13,16))
                    lastPointsTotal=pointsFromPattern+pointsFromElement+pointsFromGoal
                    if(lastPointsTotal>pointsTotal):
                        biggest=list(i)
                if(len(biggest)==0):
                    #print("the last tile cannot change anything.")
                    state.move(move)
                else:
                    state.move(biggest)
                biggest=[]
                moves=state.possible_moves2()
                move=[]
                for i in moves:
                    points2FromElement=state.calculate_element_points(state.tile_base2)
                    points2FromPattern=state.calculate_pattern_points(state.tile_base2)
                    points2FromGoal=state.calculate_goal_points(state.tile_base2,(6,13,16))
                    points2Total=points2FromPattern+points2FromElement+points2FromGoal
                    move=list(i)
                    state.last_player2_place_tile(move)
                    points2FromElement=state.calculate_element_points(state.tile_base2)
                    points2FromPattern=state.calculate_pattern_points(state.tile_base2)
                    points2FromGoal=state.calculate_goal_points(state.tile_base2,(6,13,16))
                    lastPointsTotal=points2FromPattern+points2FromElement+points2FromGoal
                    if(lastPointsTotal>points2Total):
                        biggest=list(i)
                if(len(biggest)==0):
                    #print("the last tile cannot change anything.")
                    state.move(move)
                else:
                    state.move(biggest)
                #print(state.to_play)
            else:
                #print("player1 is thinking...")
                mcts.search(searchTime)
                num_rollouts, run_time = mcts.statistics()
                rollOut.append(num_rollouts)
                print("Statistics",state.round,": ", num_rollouts, "rollouts in", run_time, "seconds")
                move = mcts.best_move()
                #print("MCTS chose move: ", move)
                state.move(move)
                mcts.move(move)
                mcts2.move(move)

                #print("player2 is thinking...")   
                mcts2.search(searchTime)
                num_rollouts, run_time = mcts2.statistics()
                rollOut.append(num_rollouts)
                print("Statistics",state.round,": ", num_rollouts, "rollouts in", run_time, "seconds")
                move2 = mcts2.best_move()
                #print("MCTS chose move: ", move2)
                state.move(move2)
                mcts2.move(move2)
                mcts.move(move2)
        data2.append(rollOut)
        pointsFromElement=0
        pointsFromPattern=0
        pointsFromElement+=state.rainbow_button(state.tile_base)
        for i in range(len(state.tile_base)):
            if(state.tile_base[i]['token']==1):
                pointsFromElement+=3
            elif(state.tile_base[i]['token']==2):
                pointsFromPattern+=3
            elif(state.tile_base[i]['token']==3):
                pointsFromPattern+=5
            elif(state.tile_base[i]['token']==4):
                pointsFromPattern+=7
        pointsFromGoal=state.calculate_goal_points(state.tile_base,(6,13,16))
        pointsTotal=pointsFromPattern+pointsFromElement+pointsFromGoal
        print('pointsTotal',pointsTotal)
        #print('player1 draw %d points from element, %d points from pattern, %d points from goal'%(pointsFromElement,pointsFromPattern,pointsFromGoal))
        #print('player1 gains',pointsTotal)
        buttonToken+=pointsFromElement
        catToken+=pointsFromPattern
        designGoal+=pointsFromGoal
        total+=(pointsFromElement+pointsFromPattern+pointsFromGoal)

        wholeTotal.append((pointsFromElement+pointsFromPattern+pointsFromGoal))
        maximumTotal=max(wholeTotal)
        minimumTotal=min(wholeTotal)
        roundCount.append(pointsFromElement)
        roundCount.append(pointsFromPattern)
        roundCount.append(pointsFromGoal)
        roundCount.append(pointsTotal)
 
        points2FromElement=0
        points2FromPattern=0
        points2FromElement+=state.rainbow_button(state.tile_base2)
        for i in range(len(state.tile_base2)):
            if(state.tile_base2[i]['token']==1):
                points2FromElement+=3
            elif(state.tile_base2[i]['token']==2):
                points2FromPattern+=3
            elif(state.tile_base2[i]['token']==3):
                points2FromPattern+=5
            elif(state.tile_base2[i]['token']==4):
                points2FromPattern+=7
        points2FromGoal=state.calculate_goal_points(state.tile_base2,(6,13,16))
        points2Total=points2FromPattern+points2FromElement+points2FromGoal
        print('points2Total',points2Total)
        #print('player2 draw %d points from element, %d points from pattern, %d points from goal'%(points2FromElement,points2FromPattern,points2FromGoal))
        #print('player1 gains',pointsTotal)
        #print('player2 gains',points2Total)
        buttonToken2+=points2FromElement
        catToken2+=points2FromPattern
        designGoal2+=points2FromGoal
        total2+=(points2FromElement+points2FromPattern+points2FromGoal)
        wholeTotal2.append((points2FromElement+points2FromPattern+points2FromGoal))
        maximumTotal2=max(wholeTotal2)
        minimumTotal2=min(wholeTotal2)
        roundCount.append(points2FromElement)
        roundCount.append(points2FromPattern)
        roundCount.append(points2FromGoal)
        roundCount.append(points2Total)
        if(state.winner==1):
            player1Win+=1
            #print('player1 win')
        elif(state.winner==2):
            player2Win+=1
            #print('player2 win')
        else:
            tie+=1
        data.append(roundCount)
        

# random vs. MCTS
    """workbook = xlsxwriter.Workbook('randomVsMCTSExp(08.12).xlsx')
    wb = openpyxl.load_workbook('randomVsMCTS3.xlsx', data_only=True)

    s1 = wb.create_sheet('worksheet 1')     # add worksheet 1
    s2 = wb.create_sheet('worksheet 2')     # add worksheet 2
    data = []
    data2 = []
    dataName=['','player1','','','','player2']
    data.append(dataName)
    dataName=['','button token','cat token','goal tile','total points','button token','cat token','goal tile','total points']
    data.append(dataName)
    data2Name=[]
    data2Name.append('')
    for i in range(21):
        data2Name.append('round'+str(i+1))
    data2.append(data2Name)
    for i in range(roundTime):
        print("round",i+1,'...')
        state=GameState()
        state.initial_setting()
        state.cat_setting()
        state.set_goal_tile(state.tile_base)
        state.set_goal_tile(state.tile_base2)
        state.place_goal_tile(state.tile_base,(0,1,2),(6,13,16))
        state.place_goal_tile(state.tile_base2,(0,1,2),(6,13,16))
        mcts2=UctMctsAgent2(state)
        
        rollOut=[]
        roundCount=[]
        roundCount.append('round'+str(i+1))
        rollOut.append('round'+str(i+1))
        
        while state.round<=22:
            if(state.round==22):
                moves=state.possible_moves()
                move=random.choice(moves)
                state.move(move)
                mcts2.move(move)

                biggest=[]
                moves=state.possible_moves2()
                move=[]
                for i in moves:
                    points2FromElement=state.calculate_element_points(state.tile_base2)
                    points2FromPattern=state.calculate_pattern_points(state.tile_base2)
                    points2FromGoal=state.calculate_goal_points(state.tile_base2,(6,13,16))
                    points2Total=points2FromPattern+points2FromElement+points2FromGoal
                    move=list(i)
                    state.last_player2_place_tile(move)
                    points2FromElement=state.calculate_element_points(state.tile_base2)
                    points2FromPattern=state.calculate_pattern_points(state.tile_base2)
                    points2FromGoal=state.calculate_goal_points(state.tile_base2,(6,13,16))
                    lastPointsTotal=points2FromPattern+points2FromElement+points2FromGoal
                    if(lastPointsTotal>points2Total):
                        biggest=list(i)
                if(len(biggest)==0):
                    state.move(move)
                else:
                    state.move(biggest)
            else:
                moves=state.possible_moves()
                move=random.choice(moves)
                state.move(move)
                mcts2.move(move)

                mcts2.search(searchTime)
                num_rollouts, run_time = mcts2.statistics()
                rollOut.append(num_rollouts)
                print("Statistics",state.round,": ", num_rollouts, "rollouts in", run_time, "seconds")
                move2 = mcts2.best_move()
                state.move(move2)
                mcts2.move(move2)
        data2.append(rollOut)
        pointsFromElement=0
        pointsFromPattern=0
        pointsFromElement+=state.rainbow_button(state.tile_base)
        for i in range(len(state.tile_base)):
            if(state.tile_base[i]['token']==1):
                pointsFromElement+=3
            elif(state.tile_base[i]['token']==2):
                pointsFromPattern+=3
            elif(state.tile_base[i]['token']==3):
                pointsFromPattern+=5
            elif(state.tile_base[i]['token']==4):
                pointsFromPattern+=7
        pointsFromGoal=state.calculate_goal_points(state.tile_base,(6,13,16))
        pointsTotal=pointsFromPattern+pointsFromElement+pointsFromGoal
        print('pointsTotal',pointsTotal)
        buttonToken+=pointsFromElement
        catToken+=pointsFromPattern
        designGoal+=pointsFromGoal
        total+=(pointsFromElement+pointsFromPattern+pointsFromGoal)
        wholeTotal.append((pointsFromElement+pointsFromPattern+pointsFromGoal))
        maximumTotal=max(wholeTotal)
        minimumTotal=min(wholeTotal)
        roundCount.append(pointsFromElement)
        roundCount.append(pointsFromPattern)
        roundCount.append(pointsFromGoal)
        roundCount.append(pointsTotal)
 
        points2FromElement=0
        points2FromPattern=0
        points2FromElement+=state.rainbow_button(state.tile_base2)
        for i in range(len(state.tile_base2)):
            if(state.tile_base2[i]['token']==1):
                points2FromElement+=3
            elif(state.tile_base2[i]['token']==2):
                points2FromPattern+=3
            elif(state.tile_base2[i]['token']==3):
                points2FromPattern+=5
            elif(state.tile_base2[i]['token']==4):
                points2FromPattern+=7
        points2FromGoal=state.calculate_goal_points(state.tile_base2,(6,13,16))
        points2Total=points2FromPattern+points2FromElement+points2FromGoal
        print('points2Total',points2Total)
        buttonToken2+=points2FromElement
        catToken2+=points2FromPattern
        designGoal2+=points2FromGoal
        total2+=(points2FromElement+points2FromPattern+points2FromGoal)
        wholeTotal2.append((points2FromElement+points2FromPattern+points2FromGoal))
        maximumTotal2=max(wholeTotal2)
        minimumTotal2=min(wholeTotal2)
        roundCount.append(points2FromElement)
        roundCount.append(points2FromPattern)
        roundCount.append(points2FromGoal)
        roundCount.append(points2Total)
        if(state.winner==1):
            player1Win+=1
        elif(state.winner==2):
            player2Win+=1
        else:
            tie+=1
        data.append(roundCount)"""
    # Workbook() takes one, non-optional, argument
    # which is the filename that we want to create.
    
    """for i in data:
        s1.append(i)
    for i in data2:
        s2.append(i)
    wb.save('newRandomVsMCTS(08.12).xlsx')"""

#MCTS vs. random

    """workbook = xlsxwriter.Workbook('MCTSVsRandomC05.xlsx')
    for i in range(roundTime):
        print("round",i+1,'...')
        state=GameState()
        state.initial_setting()
        state.cat_setting()
        state.set_goal_tile(state.tile_base)
        state.set_goal_tile(state.tile_base2)
        state.place_goal_tile(state.tile_base,(0,1,2),(6,13,16))
        state.place_goal_tile(state.tile_base2,(0,1,2),(6,13,16))

        mcts=UctMctsAgent(state)
        #mcts2=UctMctsAgent2(state)

        

        while state.round<=22:
            if(state.round==22):
                biggest=[]
                moves=state.possible_moves()
                move=[]
                for i in moves:
                    pointsFromElement=state.calculate_element_points(state.tile_base)
                    pointsFromPattern=state.calculate_pattern_points(state.tile_base)
                    pointsFromGoal=state.calculate_goal_points(state.tile_base,(6,13,16))
                    pointsTotal=pointsFromPattern+pointsFromElement+pointsFromGoal
                    move=list(i)
                    state.last_player1_place_tile(move)
                    pointsFromElement=state.calculate_element_points(state.tile_base)
                    pointsFromPattern=state.calculate_pattern_points(state.tile_base)
                    pointsFromGoal=state.calculate_goal_points(state.tile_base,(6,13,16))
                    lastPointsTotal=pointsFromPattern+pointsFromElement+pointsFromGoal
                    if(lastPointsTotal>pointsTotal):
                        biggest=list(i)
                if(len(biggest)==0):
                    #print("the last tile cannot change anything.")
                    state.move(move)
                else:
                    state.move(biggest)
                
                moves=state.possible_moves2()
                move=random.choice(moves)
                state.move(move)
                mcts.move(move)
            else:
                print("player1 is thinking...")   
                mcts.search(searchTime)
                num_rollouts, run_time = mcts.statistics()
                print("Statistics",state.round,": ", num_rollouts, "rollouts in", run_time, "seconds")
                move = mcts.best_move()
                print("MCTS chose move: ", move)
                state.move(move)
                mcts.move(move)

                moves=state.possible_moves2()
                move=random.choice(moves)
                state.move(move)
                mcts.move(move)
    print('tile_base=')
    for i in range(len(state.tile_base)):
        print(state.tile_base[i])
    for i in range(len(state.edgeTile)):
        print(state.edgeTile[i])
    print('tile_base2=')
    for i in range(len(state.tile_base2)):
        print(state.tile_base2[i])"""

    for i in data:
        s1.append(i)
    for i in data2:
        s2.append(i)
    wb.save('newMCTSVsMCTS.xlsx')
    # The workbook object is then used to add new
    # worksheet via the add_worksheet() method.
    worksheet = workbook.add_worksheet()

    # Use the worksheet object to write
    # data via the write() method.
    worksheet.write('A1', '')
    worksheet.write('B1','WINS')
    worksheet.write('C1', 'button')
    worksheet.write('D1', 'cat')
    worksheet.write('E1', 'goal tile')
    worksheet.write('F1', 'total')
    worksheet.write('G1','WN')#winning rate
    worksheet.write('H1','max')
    worksheet.write('I1','min')
    worksheet.write('A2', 'player1')
    worksheet.write('A3', 'player2')
    worksheet.write('J5','round')
    worksheet.write('J6','C')
    worksheet.write('J7','search')

    worksheet.write('B2',player1Win)
    worksheet.write('C2', (buttonToken/roundTime))
    worksheet.write('D2', (catToken/roundTime))
    worksheet.write('E2', (designGoal/roundTime))
    worksheet.write('F2', (total/roundTime))
    worksheet.write('G2',(player1Win/roundTime))
    worksheet.write('H2',maximumTotal)
    worksheet.write('I2',minimumTotal)
    worksheet.write('B3',player2Win)
    worksheet.write('C3', (buttonToken2/roundTime))
    worksheet.write('D3', (catToken2/roundTime))
    worksheet.write('E3', (designGoal2/roundTime))
    worksheet.write('F3', (total2/roundTime))
    worksheet.write('G3',(player2Win/roundTime))
    worksheet.write('H3',maximumTotal2)
    worksheet.write('I3',minimumTotal2)

    worksheet.write('G4',(tie/roundTime))
    worksheet.write('K5',roundTime)
    worksheet.write('K6','05')
    worksheet.write('K7',searchTime)

    workbook.close()
    print('player1:')
    print('WINS:',player1Win)
    print('button:',(buttonToken/roundTime))
    print('cat:',(catToken/roundTime))
    print('goal tile:',(designGoal/roundTime))
    print('total:',(total/roundTime))
    print('win rate:',(player1Win/roundTime))
    print('maximumPoint:',maximumTotal)
    print('minimumPoint:',minimumTotal)
    print('player2:')
    print('WINS:',player2Win)
    print('button:',(buttonToken2/roundTime))
    print('cat:',(catToken2/roundTime))
    print('goal tile:',(designGoal2/roundTime))
    print('total:',(total2/roundTime))
    print('win rate:',(player2Win/roundTime))
    print('maximumPoint:',maximumTotal2)
    print('minimumPoint:',minimumTotal2)
    print('round:',roundTime)
    print('C:05')
    print('searchTime:',searchTime)
